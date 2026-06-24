"""UCDP armed-conflict world maps — download, aggregate, join, render.

Pipeline (all backend-aware via :mod:`conflict.storage`):

1. ``download_ucdp_aggregate`` — fetch the UCDP GED bulk CSV (the Georeferenced
   Event Dataset; the public API now needs a token, the bulk zip does not),
   aggregate the latest year by country into the 5 metrics, and cache the small
   aggregate JSON (so re-runs skip the 250 MB parse).
2. ``build_conflict_map`` — join the aggregate onto Natural Earth country
   geometry (with a small UCDP→NE name alias map), normalise intensity by the
   geometry's ``POP_EST``, and render a MapLibre world choropleth with a metric
   dropdown ("dark = worse") + a provenance footer.

Metrics (UCDP-derivable subset):
  events    — conflict events that year (GED rows)
  deaths    — fatalities that year (Σ ``best``)
  civilian  — violence-against-civilians events (``type_of_violence`` == 3)
  intensity — deaths per 100,000 population
  actors    — distinct armed groups (union of ``side_a`` / ``side_b``)
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import tempfile
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from html import escape

from . import storage as cstore

try:
    import requests
except ImportError:  # pragma: no cover
    requests = None

logger = logging.getLogger("conflict")
csv.field_size_limit(20_000_000)

UCDP_GED_URL = "https://ucdp.uu.se/downloads/ged/ged251-csv.zip"
WORLD_GEOJSON_URL = (
    "https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/"
    "geojson/ne_110m_admin_0_countries.geojson"
)
FFL_URL = "https://github.com/rlemke/fwh_conflict/blob/main/src/conflict/ffl/conflict.ffl"
USER_AGENT = "facetwork-conflict/1.0 (+https://github.com/rlemke/facetwork)"

# UCDP country names → Natural Earth NAME (the handful that differ).
COUNTRY_ALIASES = {
    "Russia (Soviet Union)": "Russia",
    "DR Congo (Zaire)": "Dem. Rep. Congo",
    "Myanmar (Burma)": "Myanmar",
    "Yemen (North Yemen)": "Yemen",
    "Central African Republic": "Central African Rep.",
    "South Sudan": "S. Sudan",
    "Cambodia (Kampuchea)": "Cambodia",
    "Madagascar (Malagasy)": "Madagascar",
    "Bosnia-Herzegovina": "Bosnia and Herz.",
    "Ivory Coast": "Côte d'Ivoire",
}

# YlOrRd ramp (fraction → colour), light → dark = worse.
RAMP = [
    [0.0, "#ffffb2"], [0.25, "#fecc5c"], [0.5, "#fd8d3c"],
    [0.75, "#f03b20"], [1.0, "#bd0026"],
]
NODATA = "#e0e0e0"


@dataclass
class Metric:
    key: str
    label: str
    fmt: str  # "count" | "rate"
    worse: str = "high"


METRICS = [
    Metric("events", "Conflict events", "count"),
    Metric("deaths", "Conflict deaths", "count"),
    Metric("civilian", "Civilian targeting", "count"),
    Metric("intensity", "Conflict intensity (deaths / 100k)", "rate"),
    Metric("actors", "Armed actor count", "count"),
    # UNHCR (joined by ISO3, not UCDP) — people displaced FROM each country.
    Metric("displaced", "Displaced population (refugees + IDPs)", "count"),
    # IDMC — new conflict displacements this year (joined by ISO3).
    Metric("new_displaced", "New displacements (conflict, this year)", "count"),
    # IPC — population in acute food insecurity Phase 3+ (joined by ISO3).
    Metric("food_insecure", "Food insecurity (IPC Phase 3+)", "count"),
]

UNHCR_URL = "https://api.unhcr.org/population/v1/population/"
# IDMC GIDD export — IDMCWSHSOLO009 is IDMC's PUBLIC client_id (used by their own
# HDX records); the bulk export is open, no registration.
IDMC_EXPORT_URL = (
    "https://helix-tools-api.idmcdb.org/external-api/gidd/displacements/displacement-export/"
    "?client_id=IDMCWSHSOLO009&release_environment=RELEASE"
)
# IPC global acute food insecurity, open CSV on HDX (no API key).
IPC_CSV_URL = (
    "https://data.humdata.org/dataset/7a7e7428-b8d7-4d2e-91d3-19100500e016/resource/"
    "6926dff7-658a-49e1-8d61-0ed8a983fbe1/download/ipc_global_national_long_latest.csv"
)


@dataclass
class ConflictMapResult:
    output_path: str
    html_path: str
    year: int
    country_count: int


# ---------------------------------------------------------------------------
# Download + aggregate.
# ---------------------------------------------------------------------------


def download_ucdp_aggregate(*, year: int | None = None, force: bool = False) -> tuple[int, dict]:
    """Return (year, {ucdp_country: {events, deaths, civilian, actors:set→count}}).

    Caches the small aggregate JSON; only re-parses the 250 MB CSV when forced
    or absent. ``year`` defaults to the latest year present in the dataset."""
    cache_key = cstore.join(cstore.cache_root(), f"aggregate-{year or 'latest'}.json")
    if not force and cstore.exists(cache_key):
        with cstore.open_read(cache_key) as f:
            blob = json.load(f)
        return blob["year"], blob["countries"]

    if requests is None:
        raise RuntimeError("requests is required to download the UCDP dataset")
    logger.info("downloading UCDP GED %s", UCDP_GED_URL)
    resp = requests.get(UCDP_GED_URL, timeout=300, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    csv_name = next(n for n in zf.namelist() if n.lower().endswith(".csv"))

    ev: Counter = Counter()
    deaths: Counter = Counter()
    civ: Counter = Counter()
    actors: dict[str, set] = defaultdict(set)
    years: Counter = Counter()
    with zf.open(csv_name) as raw:
        rows = list(csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8")))
    for row in rows:
        try:
            years[int(row["year"])] += 1
        except (ValueError, TypeError, KeyError):
            continue
    target = year or max(years)
    for row in rows:
        if row.get("year") != str(target):
            continue
        c = (row.get("country") or "").strip()
        if not c:
            continue
        ev[c] += 1
        try:
            deaths[c] += int(row.get("best") or 0)
        except ValueError:
            pass
        if row.get("type_of_violence") == "3":
            civ[c] += 1
        for s in (row.get("side_a"), row.get("side_b")):
            if s and s.strip():
                actors[c].add(s.strip())

    countries = {
        c: {"events": ev[c], "deaths": deaths[c], "civilian": civ[c], "actors": len(actors[c])}
        for c in ev
    }
    with cstore.open_write(cache_key, "w") as f:
        json.dump({"year": target, "countries": countries}, f)
    logger.info("aggregated UCDP %s: %d countries", target, len(countries))
    return target, countries


def _world_geojson() -> dict:
    cache_key = cstore.join(cstore.cache_root(), "world-countries.geojson")
    if cstore.exists(cache_key):
        with cstore.open_read(cache_key) as f:
            return json.load(f)
    if requests is None:
        raise RuntimeError("requests is required to download world geometry")
    resp = requests.get(WORLD_GEOJSON_URL, timeout=120, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    data = resp.json()
    with cstore.open_write(cache_key, "w") as f:
        json.dump(data, f)
    return data


def download_unhcr_displacement(*, year: int, force: bool = False) -> dict[str, int]:
    """Return {ISO3: refugees + IDPs displaced FROM that country} for ``year``.

    UNHCR Refugee Statistics population endpoint, aggregated by country of
    origin (``coo_iso``). Refugees who fled the country + IDPs within it = the
    country's forced-displacement burden. Cached (small)."""
    cache_key = cstore.join(cstore.cache_root(), f"unhcr-displacement-{year}.json")
    if not force and cstore.exists(cache_key):
        with cstore.open_read(cache_key) as f:
            return json.load(f)
    if requests is None:
        raise RuntimeError("requests is required to download UNHCR data")
    out: dict[str, int] = {}
    page, max_pages = 1, 1
    while page <= max_pages:
        resp = requests.get(
            UNHCR_URL, params={"year": year, "coo_all": "true", "limit": 1000, "page": page},
            timeout=120, headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
        )
        resp.raise_for_status()
        payload = resp.json()
        max_pages = int(payload.get("maxPages", 1) or 1)
        for r in payload.get("items", []):
            iso = r.get("coo_iso")
            if not iso or iso == "-":
                continue
            try:
                out[iso] = out.get(iso, 0) + int(r.get("refugees") or 0) + int(r.get("idps") or 0)
            except (TypeError, ValueError):
                pass
        page += 1
    logger.info("UNHCR displacement %s: %d origin countries", year, len(out))
    with cstore.open_write(cache_key, "w") as f:
        json.dump(out, f)
    return out


def download_idmc_new_displacements(*, year: int, force: bool = False) -> dict[str, int]:
    """Return {ISO3: new CONFLICT displacements in ``year``} from the IDMC GIDD
    bulk export (public client_id, open). "Conflict Internal Displacements" =
    new displacement events that year. Cached."""
    cache_key = cstore.join(cstore.cache_root(), f"idmc-new-displacements-{year}.json")
    if not force and cstore.exists(cache_key):
        with cstore.open_read(cache_key) as f:
            return json.load(f)
    if requests is None:
        raise RuntimeError("requests is required to download IDMC data")
    import openpyxl  # heavier dep, imported lazily
    resp = requests.get(IDMC_EXPORT_URL, params={"start_year": year, "end_year": year},
                        timeout=180, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    ci = {str(n): i for i, n in enumerate(next(rows))}
    iso_i, val_i = ci.get("ISO3"), ci.get("Conflict Internal Displacements")
    out: dict[str, int] = {}
    if iso_i is not None and val_i is not None:
        for r in rows:
            if len(r) <= max(iso_i, val_i):
                continue
            iso, v = r[iso_i], r[val_i]
            if iso and v:
                try:
                    out[iso] = out.get(iso, 0) + int(v)
                except (TypeError, ValueError):
                    pass
    logger.info("IDMC new conflict displacements %s: %d countries", year, len(out))
    with cstore.open_write(cache_key, "w") as f:
        json.dump(out, f)
    return out


def download_ipc_food_insecurity(*, force: bool = False) -> dict[str, int]:
    """Return {ISO3: population in IPC Phase 3+ (crisis or worse)} from the latest
    "current" analysis (open HDX CSV, no key). A current snapshot of acute food
    insecurity — not year-specific. Cached."""
    cache_key = cstore.join(cstore.cache_root(), "ipc-food-insecurity.json")
    if not force and cstore.exists(cache_key):
        with cstore.open_read(cache_key) as f:
            return json.load(f)
    if requests is None:
        raise RuntimeError("requests is required to download IPC data")
    resp = requests.get(IPC_CSV_URL, timeout=120, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    out: dict[str, int] = {}
    for row in csv.DictReader(io.StringIO(resp.content.decode("utf-8-sig"))):
        if row.get("Phase") == "3+" and row.get("Validity period") == "current":
            iso = row.get("Country")
            try:
                if iso:
                    out[iso] = int(row.get("Number") or 0)
            except (TypeError, ValueError):
                pass
    logger.info("IPC Phase 3+ : %d countries", len(out))
    with cstore.open_write(cache_key, "w") as f:
        json.dump(out, f)
    return out


# ---------------------------------------------------------------------------
# Build the map.
# ---------------------------------------------------------------------------


def build_conflict_map(*, year: int | None = None, force: bool = False) -> ConflictMapResult:
    """Aggregate UCDP, join onto world geometry, render the choropleth."""
    target, countries = download_ucdp_aggregate(year=year, force=force)
    displaced = download_unhcr_displacement(year=target, force=force)
    new_displaced = download_idmc_new_displacements(year=target, force=force)
    food_insecure = download_ipc_food_insecurity(force=force)
    world = _world_geojson()

    matched = 0
    for ft in world["features"]:
        props = ft["properties"]
        name = props.get("NAME", "")
        pop = props.get("POP_EST")
        # ISO3 for the UNHCR join (NE puts -99 on a few, e.g. France/Norway → ADM0_A3).
        iso = props.get("ISO_A3")
        if not iso or iso == "-99":
            iso = props.get("ADM0_A3")
        # UCDP joins by NAME (with an alias map for its historical names).
        rec = countries.get(name)
        if rec is None:
            for ucdp, ne in COUNTRY_ALIASES.items():
                if ne == name and ucdp in countries:
                    rec = countries[ucdp]
                    break

        new = {f"m_{m.key}": None for m in METRICS}
        if rec:
            matched += 1
            new["m_events"] = rec["events"]
            new["m_deaths"] = rec["deaths"]
            new["m_civilian"] = rec["civilian"]
            new["m_actors"] = rec["actors"]
            new["m_intensity"] = (
                round(rec["deaths"] / pop * 100000, 2) if pop and pop > 0 else None
            )
        new["m_displaced"] = displaced.get(iso)       # UNHCR, by ISO3
        new["m_new_displaced"] = new_displaced.get(iso)  # IDMC, by ISO3
        new["m_food_insecure"] = food_insecure.get(iso)  # IPC, by ISO3
        ft["properties"] = {"NAME": name, "POP_EST": pop, **new}

    fc = {"type": "FeatureCollection", "features": world["features"]}
    html = _render_html(fc, year=target)

    out_dir = cstore.output_root()
    geojson_path = cstore.join(out_dir, "conflict.geojson")
    html_path = cstore.join(out_dir, "index.html")
    with cstore.open_write(geojson_path, "w") as f:
        json.dump(fc, f, separators=(",", ":"))
    with cstore.open_write(html_path, "w") as f:
        f.write(html)
    return ConflictMapResult(geojson_path, html_path, target, matched)


def _attribution(year: int | None = None) -> str:
    call = "conflict.workflows.BuildConflictWorldMap"
    if year:
        call = f"{call}(year={year})"
    return (
        '<div style="position:fixed;bottom:10px;left:10px;z-index:9999;'
        "background:rgba(255,255,255,0.92);border-radius:6px;padding:6px 10px;"
        "box-shadow:0 1px 4px rgba(0,0,0,0.2);font:11px system-ui,sans-serif;color:#444;"
        'max-width:440px">Generated by Facetwork workflow '
        '<code style="background:#f0f0f0;padding:0 3px;border-radius:3px">'
        f"{escape(call)}</code> &middot; "
        f'<a href="{escape(FFL_URL)}" target="_blank" rel="noopener" '
        'style="color:#1565c0;text-decoration:none">view FFL</a></div>'
    )


def _render_html(fc: dict, *, year: int) -> str:
    data_js = json.dumps(fc, separators=(",", ":"))
    metrics_js = json.dumps([{"key": f"m_{m.key}", "label": m.label, "fmt": m.fmt} for m in METRICS])
    ramp_js = json.dumps(RAMP)
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Armed conflict by country - {year}</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link href="https://unpkg.com/maplibre-gl@4.7.1/dist/maplibre-gl.css" rel="stylesheet">
<script src="https://unpkg.com/maplibre-gl@4.7.1/dist/maplibre-gl.js"></script>
<style>
  html,body,#map{{margin:0;height:100%;width:100%;font-family:system-ui,sans-serif}}
  .panel{{position:absolute;z-index:1;background:rgba(255,255,255,.94);padding:10px 12px;
    border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.3);font-size:12px}}
  #ctl{{top:10px;left:10px;max-width:330px}}
  #ctl h3{{margin:0 0 6px;font-size:14px}} #ctl select{{font-size:13px;padding:3px;width:100%}}
  #legend{{bottom:18px;right:10px}} #legend .scale{{display:flex;margin-top:4px}}
  #legend .scale div{{display:flex;flex-direction:column;align-items:center;font-size:10px}}
  #legend .scale span{{width:34px;height:12px}}
  .maplibregl-popup-content{{max-width:300px;font-size:12px}}
  .maplibregl-popup-content h4{{margin:0 0 4px;font-size:13px}}
  table.m{{border-collapse:collapse;margin-top:4px}} table.m td{{padding:1px 6px 1px 0}}
  table.m td.v{{text-align:right}} tr.sel td{{font-weight:700}}
</style></head>
<body>
<div id="map"></div>
<div id="ctl" class="panel">
  <h3>Armed conflict by country &middot; {year}</h3>
  <select id="metric"></select>
  <div style="margin-top:5px;color:#555">Armed-conflict indicators by country ({year}), shaded
  <b>dark = worse</b>. Pick a metric above; click a country for all values. Sources: UCDP
  (events, deaths, civilian targeting, actors), UNHCR (displaced), IDMC (new displacements),
  IPC (food insecurity).</div>
</div>
<div id="legend" class="panel"><b id="lgttl"></b><div class="scale" id="lgscale"></div></div>
{_attribution(year)}
<script>
const DATA={data_js}, METRICS={metrics_js}, RAMP={ramp_js};
const fmt=(v,f)=>{{ if(v===null||v===undefined||v==='') return '—';
  if(f==='rate') return (Math.round(v*10)/10)+' /100k';
  return Math.round(v).toLocaleString(); }};
const vals=k=>DATA.features.map(f=>f.properties[k]).filter(v=>typeof v==='number'&&v>0);
function colorExpr(m){{
  const a=vals(m.key); if(!a.length) return '{NODATA}';
  let lo=Math.min(...a), hi=Math.max(...a); if(lo===hi) hi=lo+1;
  const expr=['interpolate',['linear'],['get',m.key]];
  for(const [t,c] of RAMP) expr.push(lo+(hi-lo)*t, c);
  return ['case',['==',['get',m.key],null],'{NODATA}',
          ['<',['get',m.key],1],'{NODATA}',expr];
}}
function legend(m){{
  document.getElementById('lgttl').textContent=m.label;
  const a=vals(m.key); const sc=document.getElementById('lgscale'); sc.innerHTML='';
  if(!a.length) return; let lo=Math.min(...a),hi=Math.max(...a);
  RAMP.forEach(([t,c])=>{{ const d=document.createElement('div');
    d.innerHTML=`<span style="background:${{c}}"></span>${{fmt(lo+(hi-lo)*t,m.fmt)}}`; sc.appendChild(d); }});
}}
const map=new maplibregl.Map({{container:'map',style:{{version:8,
  sources:{{bm:{{type:'raster',tiles:['https://a.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}.png','https://b.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}.png'],tileSize:256,attribution:'&copy; OpenStreetMap &copy; CARTO &middot; UCDP'}}}},
  layers:[{{id:'bm',type:'raster',source:'bm'}}]}},center:[14,18],zoom:1.4}});
map.addControl(new maplibregl.NavigationControl());
const sel=document.getElementById('metric');
METRICS.forEach((m,i)=>{{const o=document.createElement('option');o.value=i;o.textContent=m.label;sel.appendChild(o);}});
let cur=METRICS[0];
map.on('load',()=>{{
  map.addSource('c',{{type:'geojson',data:DATA}});
  map.addLayer({{id:'fill',type:'fill',source:'c',paint:{{'fill-color':colorExpr(cur),'fill-opacity':0.82}}}});
  map.addLayer({{id:'line',type:'line',source:'c',paint:{{'line-color':'#888','line-width':0.3}}}});
  legend(cur);
  sel.onchange=()=>{{cur=METRICS[+sel.value];map.setPaintProperty('fill','fill-color',colorExpr(cur));legend(cur);}};
  map.on('click','fill',e=>{{const p=e.features[0].properties||{{}};
    let rows=''; for(const m of METRICS){{ const v=p[m.key];
      rows+=`<tr class="${{m.key===cur.key?'sel':''}}"><td>${{m.label}}</td><td class="v">${{fmt(v,m.fmt)}}</td></tr>`; }}
    new maplibregl.Popup({{closeButton:true,maxWidth:'300px'}}).setLngLat(e.lngLat)
      .setHTML(`<h4>${{p.NAME||'Country'}}</h4><table class="m">${{rows}}</table>`).addTo(map);}});
  map.on('mouseenter','fill',()=>map.getCanvas().style.cursor='pointer');
  map.on('mouseleave','fill',()=>map.getCanvas().style.cursor='');
}});
</script></body></html>"""
